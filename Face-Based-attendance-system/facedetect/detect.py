#!/usr/bin/python
import numpy as np
import cv2
import pandas as pd
import mysql.connector
import sys
import csv
import os
from pandas import read_csv
from tkinter import *
from tkinter import filedialog
import openpyxl
from PIL import Image;


db=mysql.connector.connect(host='127.0.0.1',user='root',passwd='1810',db='student' )
cur=db.cursor()


#just create a column in the csv file
def mark():
    mtext=ment.get()
    wb = openpyxl.load_workbook('attendance.xlsx')
    sheet = wb['Sheet1']
    columns = sheet.max_column+1
    sheet.insert_cols(columns,1)
    sheet.cell(row=1, column=columns).value =mtext
    wb.save("C:/xampp1/htdocs/facedetect/attendance.xlsx")

#detecting faces and marking attendance
def open_file():
    result=  filedialog.askopenfile(initialdir = "/xampp1/htdocs/facedetect",title = "Select file",filetypes = (("jpeg files","*.jpg"),("all files","*.*")))
    filename = os.path.abspath(result.name)

    face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')

    rec = cv2.face.LBPHFaceRecognizer_create()
    rec.read("trainer/trainer.yml")
    my_list=[]


    img = cv2.imread(filename)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    # noise removal using iterative bilateral filters(removing noise and preserving edges

    gray = cv2.bilateralFilter(gray, 11, 17, 17)
    faces = face_cascade.detectMultiScale(gray, 1.5, 5)
    for (x, y, w, h) in faces:
        cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 255), 2)
        id, conf = rec.predict(gray[y:y + h, x:x + w])

        if conf > 75:
            my_list.append(id)

    cv2.imshow("detected faces", img)
    cv2.waitKey(0)


    df1 = pd.read_excel('attendance.xlsx')

    row, column = df1.shape
    df1.loc[:, df1.columns[column - 1]] = 0
    for i in my_list:
        df1.loc[df1['ID'] == i, df1.columns[column - 1]] = 1
    print(df1)
    df1.to_csv(r'C:/xampp1/htdocs/facedetect/markedattendance.csv')


#training with dataset and training the model and saving trained data in a yml file
def train():
    recognizer = cv2.face.LBPHFaceRecognizer_create()
    detector = cv2.CascadeClassifier("haarcascade_frontalface_default.xml");
    path = 'C:/xampp1/htdocs/facedetect/dataset'

    def getImagesAndLabels(path):
        # get the path of all the files in the folder
        imagePaths = [os.path.join(path, f) for f in os.listdir(path)]
        # create empth face list
        faceSamples = []
        # create empty ID list
        Ids = []
        # now looping through all the image paths and loading the Ids and the images
        for imagePath in imagePaths:
            # loading the image and converting it to gray scale
            pilImage = Image.open(imagePath).convert('L')
            # Now we are converting the PIL image into numpy array
            imageNp = np.array(pilImage, 'uint8')
            # getting the Id from the image
            Id = int(os.path.split(imagePath)[-1].split(".")[1])
            # extract the face from the training image sample
            faces = detector.detectMultiScale(imageNp)
            # If a face is there then append that in the list as well as Id of it
            for (x, y, w, h) in faces:
                faceSamples.append(imageNp[y:y + h, x:x + w])
                Ids.append(Id)
        return faceSamples, Ids

    faces, Ids = getImagesAndLabels('dataSet')
    recognizer.train(faces, np.array(Ids))
    print("Successfully trained")
    recognizer.write('trainer/trainer.yml')
#create the attendance sheet
def create():
    QUERY = 'SELECT * FROM studentdetails;'
    cur.execute(QUERY)
    result = cur.fetchall()
    with open('dbdump01.csv', 'w', newline='') as outcsv:
        writer = csv.DictWriter(outcsv, fieldnames=["ID", "Name", "sem","section"])
        writer.writeheader()
        c = csv.writer(outcsv,quoting=csv.QUOTE_ALL)
        for x in result:
            c.writerow(x)

    read_file = pd.read_csv('dbdump01.csv')
    read_file.to_excel(r'C:\xampp1\htdocs\facedetect\attendance.xlsx', index=None, header=True)


#predict student attendance
def pred():
    os.system('python predict.py')

root = Tk()
ment=StringVar()


label_frame = LabelFrame(root, text='select an action')
label_frame.pack(expand='yes', fill='both')
#training model after registering
button1= Button(root, text="train system with all registered faces",command=train)
button1.place(x=120, y=40)
#running attendance system
button = Button(label_frame, text="select an image an run face based attendance", command=open_file)
button.place(x=120, y=250)
#create attendance sheet with registered users
button2 = Button(label_frame, text="create  attendance sheet",command=create)
button2.place(x=120, y=80)
#entering attendance date
btn2 = Button(root, text='please enter date',command=mark)
btn2.place(x=150, y=170)
mentry=Entry(label_frame,textvariable=ment)
mentry.place(x=150,y=130)
button3= Button(label_frame, text="predict student attendance",command=pred)
button3.place(x=150, y=300)



root.geometry("400x400")
root.mainloop()



